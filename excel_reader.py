import pandas as pd
from quadrant import Quadrant
import openpyxl
from pathlib import Path
from openpyxl import Workbook, load_workbook


class ExcelReader:
    def __init__(self, forecast_excel_file_path,planning_excel_file_path,  sheet_names_list = []):
        self.forecast_excel_file_path = forecast_excel_file_path
        self.planning_excel_file_path = planning_excel_file_path
        self.sheet_names_list = sheet_names_list

        # project specific
        self.forecast_df = self.create_df_from_excel(path = self.forecast_excel_file_path, sheet_name = "forecast 2025")
        self.quadrants_df = self.create_df_from_excel(path = self.planning_excel_file_path, sheet_name = "quadrants")


    # GENERAL METHODS 
    def create_df_from_excel(self, sheet_name, path):     #returns a dictionary with all the info from the info sheet
        excel_df = pd.read_excel(path, sheet_name=sheet_name, header=0, dtype=str)
        return excel_df
    
    def unmerge_and_fill_values(self, file_path_merged, file_path_unmerged, sheet_name):
        # Load the workbook and select the sheet
        workbook = load_workbook(file_path_merged)
        sheet = workbook[sheet_name]

        # Loop through all merged cells
        for merged_range in list(sheet.merged_cells.ranges):  # Use list() to avoid modifying while iterating
            # Get the top-left cell of the merged range
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            top_left_cell = sheet.cell(min_row, min_col)
            value = top_left_cell.value  # Get the value of the merged cell

            # Fill all cells in the merged range with the same value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row, col).value = value

            # Unmerge the cells
            sheet.unmerge_cells(str(merged_range))

        # Save the workbook
        workbook.save(file_path_unmerged)
        print(f"Successfully unmerged cells and filled values in {sheet_name}.")
    
    #PROJECT SPECIFIC METHODS
    def get_product_count(self, product_type, product_size, product_force, month ):
        # Filter rows by product type, size, and force
        filtered_df = self.forecast_df[
            (self.forecast_df.iloc[:, 1].str.contains(product_size, na=False)) &
            (self.forecast_df.iloc[:, 2].str.contains(product_force, na=False)) &
            (self.forecast_df.iloc[:, 3].str.contains(product_type, na=False))
        ]
        
        # Find the column that matches the given month
        matching_columns = [col for col in self.forecast_df.columns if month in col]
        if not matching_columns:
            raise ValueError(f"Month '{month}' not found in forecast data headers.")
        
        # Use the first matching column (in case of duplicates, which should ideally not happen)
        month_column = matching_columns[0]
        
        # Retrieve the product count for the specified month
        if not filtered_df.empty:
            # Convert the month column values to float and sum them
            product_count = filtered_df[month_column].astype(float).sum() if not filtered_df.empty else 0
            return product_count
        else:
            print(" no matching data found")
            return None

        # Read the Excel file into a DataFrame
        sheet_name = "slides"
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0)

        # Create a list of Question objects
        questions = []
        for _, row in df.iterrows():
            question = Question(
                variable_label=row['variable_label'],
                question=row['question'],
                question_type=row['type'],
                recode=row['recode'],
                shell=row['shell'],
                filter=row['filter'],
                routing=row['routing'],
                slide_title = row['slide titel'],
                amount_of_respondents = row['n'],
            )
            questions.append(question)

        return questions

    def get_quadrant_info(self, quadrant_id):
        # Filter the dataframe for the specified quadrant ID
        print(self.quadrants_df)
        quadrant_row = self.quadrants_df[self.quadrants_df['id'] == quadrant_id]

        if quadrant_row.empty:
            raise ValueError(f"No quadrant found with ID: {quadrant_id}")

        # Extract the details from the row
        quadrant_data = quadrant_row.iloc[0]

        # Create and return a Quadrant object using the extracted data
        return Quadrant(
            id=quadrant_data['id'],
            product=quadrant_data['product'],
            component=quadrant_data['component'],
            hfile=quadrant_data['hfile'],
            bewerkings_orde=int(quadrant_data['bewerkings_orde']),
            components_per_quadrant=int(quadrant_data['components_per_quadrant']),
            loading_time=float(quadrant_data['loading_time']),
            machine_time=float(quadrant_data['machine_time']),
            unloading_time=float(quadrant_data['unloading_time']),
            opmerkingen=quadrant_data['opmerkingen']
        )
    
    def get_quadrants_per_product(self,product_type, product_size, product_force):
        product = f"{product_type}_{product_force}_0{product_size}"

        #filter in quadrants_df op product_size, product_force
        filtered_quadrants_df = self.quadrants_df[self.quadrants_df['product'] == product]
        
        # Extract unique Quadrant IDs
        quadrant_ids = filtered_quadrants_df['id'].unique()

        # Create a list to store the quadrants
        quadrants = []

        # Generate quadrant info for each ID
        for quadrant_id in quadrant_ids:
            quadrant_info = self.get_quadrant_info(quadrant_id)
            quadrants.append(quadrant_info)

        # Return the list of quadrants
        return quadrants
    
    def get_quadrants_df_per_product(self,product_type, product_size, product_force):
        
        if(product_size == "120"):
            product = f"{product_type}_{product_force}_{product_size}"
        else:
            product = f"{product_type}_{product_force}_0{product_size}"

        #filter in quadrants_df op product_size, product_force
        filtered_quadrants_df = self.quadrants_df[self.quadrants_df['product'] == product]
        
        return filtered_quadrants_df 
    
    def get_quadrants_df(self, product_types, product_sizes, product_forces, months):

        # Initialize an empty DataFrame to hold all the combined data
        quadrants_df = pd.DataFrame()      

        length = len(product_types)
        for i in range(length):
            product_type = product_types[i]
            product_size = product_sizes[i]
            product_force = product_forces[i]
            month = months[i]
            
            quadrants_per_product_df = self.get_quadrants_df_per_product(product_type, product_size, product_force)
            product_count = self.get_product_count(product_type, product_size, product_force, month )
            product_count = int(product_count)

            # Duplicate the quadrants_df based on the product count
            repeated_quadrants_df = pd.concat([quadrants_per_product_df] * product_count, ignore_index=True)

            # Append the repeated data to the combined DataFrame
            quadrants_df = pd.concat([quadrants_df, repeated_quadrants_df], ignore_index=True)

        return quadrants_df

    def create_excel_tab_from_df_old(self, excel_path, tab_name, df):
        excel_file = Path(excel_path)
        
        # If the file doesn't exist, create and save a new workbook
        if not excel_file.exists():
            workbook = Workbook()
            # We keep the default sheet to avoid "At least one sheet must be visible" errors
            workbook.save(excel_path)
        
        # Load the existing (or newly created) workbook
        workbook = load_workbook(excel_path)

        # Check if the tab already exists
        if tab_name in workbook.sheetnames:
            raise ValueError(f"The tab '{tab_name}' already exists in the Excel file.")

        # Write the DataFrame to a new sheet in append mode
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            writer.book = workbook
            writer.sheets = {ws.title: ws for ws in workbook.worksheets}
            df.to_excel(writer, sheet_name=tab_name, index=False)
            writer.close()
        
    def create_excel_tab_from_df_old2(self, excel_path, tab_name, df):
        excel_file = Path(excel_path)
        
        # If the file doesn't exist, create and save a new workbook
        if not excel_file.exists():
            workbook = Workbook()
            workbook.save(excel_path)
        
        # Load the existing workbook
        workbook = load_workbook(excel_path)

        if tab_name in workbook.sheetnames:
            raise ValueError(f"The tab '{tab_name}' already exists in {excel_path}.")

        # Use ExcelWriter in 'append' mode
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            # Assign the loaded workbook to the underlying engine
            writer.engine.book = workbook
            
            # Let writer know about existing sheets
            writer.sheets = {ws.title: ws for ws in writer.engine.book.worksheets}
            
            # Write DataFrame to the new sheet
            df.to_excel(writer, sheet_name=tab_name, index=False)
    
    def create_excel_tab_from_df(self, excel_path, sheet, df):
        print(type(df))
        if not isinstance(df, pd.DataFrame):
            raise ValueError("The provided df is not a pandas DataFrame")
        df.to_excel(excel_path,sheet_name=sheet)
